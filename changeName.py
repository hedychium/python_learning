# encoding=utf-8

"""
__project_ = 'py-worksapce'
__file_name__ = 'changeName'
__author__ = 'li.jiang'
__time__ = '2019/6/10 16:21'
__product_name = PyCharm
# Keep calm and carry on 
"""

import pandas as pd
import openpyxl
source_dir= './Data/2017.4.21-稳健1号-穿透.xlsx'
date ='20171010'

#修改日期,合并单元格
workbook = openpyxl.load_workbook(source_dir)
sheet_name = workbook.sheetnames[0]


# data = pd.read_excel(source_dir, sheet_name=0, header=None)
# print(data)
# data.iloc[2,0] = '估值日期：%s' % date
# print(data.iloc[2,0])
# pd.DataFrame(data).to_excel(source_dir,'utf8', sheet_name, index=False,
#                             header=None, engine='openpyxl')

workbook = openpyxl.load_workbook(source_dir)
sheet_name = workbook.sheetnames[0]
worksheet = workbook[sheet_name]
worksheet.merge_cells('A3:B3')
worksheet
workbook.save(source_dir)
workbook.close()